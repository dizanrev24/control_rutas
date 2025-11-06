"""from django.shortcuts import render

Vistas para el módulo de productos.

Gestiona el catálogo de productos (CRUD completo).# Create your views here.

"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from io import BytesIO
from decimal import Decimal, InvalidOperation
import logging
from openpyxl import Workbook, load_workbook
from .models import Categoria, Producto
from .forms import ProductoForm, ProductoFiltroForm

logger = logging.getLogger(__name__)


@login_required
def producto_listar(request):
    """
    Lista todos los productos con filtros y paginación.
    
    Accesible para: Admin y Secretaría
    
    Filtros:
    - buscar: Nombre, código o categoría
    - activo: Activos/Inactivos/Todos
    - categoria: Categoría específica
    
    Args:
        request: HttpRequest
    
    Returns:
        HttpResponse: Template con lista de productos
    """
    # Validar permisos
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para gestionar productos.')
        return redirect('home')
    
    # Obtener todos los productos
    productos = Producto.objects.select_related('categoria').all().order_by('-estado', 'nombre')
    
    # Aplicar filtros
    filtro_form = ProductoFiltroForm(request.GET)
    
    if filtro_form.is_valid():
        buscar = filtro_form.cleaned_data.get('buscar')
        if buscar:
            productos = productos.filter(
                Q(nombre__icontains=buscar) |
                Q(descripcion__icontains=buscar) |
                Q(categoria__nombre__icontains=buscar)
            )
        
        estado = filtro_form.cleaned_data.get('estado')
        if estado:
            productos = productos.filter(estado=estado)
    
    # Paginación
    paginator = Paginator(productos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'filtro_form': filtro_form,
        'total_productos': productos.count(),
    }
    
    return render(request, 'productos/producto_list.html', context)


@login_required
def producto_crear(request):
    """
    Crea un nuevo producto.
    
    Accesible para: Admin y Secretaría
    
    Args:
        request: HttpRequest
    
    Returns:
        HttpResponse: Formulario o redirección
    """
    # Validar permisos
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para crear productos.')
        return redirect('home')
    
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        
        if form.is_valid():
            producto = form.save()
            messages.success(
                request,
                f'Producto "{producto.nombre}" creado exitosamente.'
            )
            return redirect('producto_listar')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ProductoForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Producto',
        'accion': 'Crear',
    }
    
    return render(request, 'productos/producto_form.html', context)


@login_required
def producto_editar(request, pk):
    """
    Edita un producto existente.
    
    Accesible para: Admin y Secretaría
    
    Args:
        request: HttpRequest
        pk: ID del producto
    
    Returns:
        HttpResponse: Formulario o redirección
    """
    # Validar permisos
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para editar productos.')
        return redirect('home')
    
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        
        if form.is_valid():
            producto = form.save()
            messages.success(
                request,
                f'Producto "{producto.nombre}" actualizado exitosamente.'
            )
            return redirect('producto_listar')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ProductoForm(instance=producto)
    
    context = {
        'form': form,
        'producto': producto,
        'titulo': f'Editar Producto: {producto.nombre}',
        'accion': 'Actualizar',
    }
    
    return render(request, 'productos/producto_form.html', context)


@login_required
def producto_eliminar(request, pk):
    """
    Desactiva un producto (soft delete).
    
    Accesible para: Admin y Secretaría
    
    Args:
        request: HttpRequest
        pk: ID del producto
    
    Returns:
        HttpResponse: Redirección
    """
    # Validar permisos
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para eliminar productos.')
        return redirect('home')
    
    producto = get_object_or_404(Producto, pk=pk)
    
    if producto.estado == 'activo':
        producto.estado = 'inactivo'
        producto.save()
        messages.success(
            request,
            f'Producto "{producto.nombre}" desactivado correctamente.'
        )
    else:
        messages.info(request, 'El producto ya estaba desactivado.')
    
    return redirect('producto_listar')


@login_required
def producto_activar(request, pk):
    """
    Activa un producto previamente desactivado.
    
    Accesible para: Admin y Secretaría
    
    Args:
        request: HttpRequest
        pk: ID del producto
    
    Returns:
        HttpResponse: Redirección
    """
    # Validar permisos
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para activar productos.')
        return redirect('home')
    
    producto = get_object_or_404(Producto, pk=pk)
    
    if producto.estado == 'inactivo':
        producto.estado = 'activo'
        producto.save()
        messages.success(
            request,
            f'Producto "{producto.nombre}" activado correctamente.'
        )
    else:
        messages.info(request, 'El producto ya estaba activo.')
    
    return redirect('producto_listar')


def categoria_lista(request):
    """
    Lista todas las categorías con búsqueda
    """
    query = request.GET.get('q', '')
    
    categorias = Categoria.objects.all()
    
    if query:
        categorias = categorias.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query)
        )
    
    # Contar productos por categoría
    for categoria in categorias:
        categoria.total_productos = categoria.productos.count()
    
    context = {
        'categorias': categorias,
        'query': query,
        'total': categorias.count(),
    }
    
    return render(request, 'productos/categoria_lista.html', context)


def categoria_crear(request):
    """
    Crea una nueva categoría
    """
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        # Validaciones
        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('categoria_crear')
        
        # Verificar si ya existe
        if Categoria.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, f'Ya existe una categoría con el nombre "{nombre}".')
            return redirect('categoria_crear')
        
        try:
            # Crear la categoría
            categoria = Categoria.objects.create(
                nombre=nombre,
                descripcion=descripcion
            )
            messages.success(request, f'Categoría "{categoria.nombre}" creada exitosamente.')
            return redirect('categoria_lista')
            
        except Exception as e:
            messages.error(request, f'Error al crear la categoría: {str(e)}')
            return redirect('categoria_crear')
    
    return render(request, 'productos/categoria_form.html')


def categoria_editar(request, pk):
    """
    Edita una categoría existente
    """
    categoria = get_object_or_404(Categoria, pk=pk)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        # Validaciones
        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('categoria_editar', pk=pk)
        
        # Verificar duplicados (excluyendo la categoría actual)
        if Categoria.objects.filter(nombre__iexact=nombre).exclude(pk=pk).exists():
            messages.error(request, f'Ya existe otra categoría con el nombre "{nombre}".')
            return redirect('categoria_editar', pk=pk)
        
        try:
            categoria.nombre = nombre
            categoria.descripcion = descripcion
            categoria.save()
            
            messages.success(request, f'Categoría "{categoria.nombre}" actualizada exitosamente.')
            return redirect('categoria_lista')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar la categoría: {str(e)}')
            return redirect('categoria_editar', pk=pk)
    
    context = {
        'categoria': categoria,
        'editando': True,
    }
    
    return render(request, 'productos/categoria_form.html', context)


def categoria_eliminar(request, pk):
    """
    Elimina una categoría si no tiene productos asociados
    """
    categoria = get_object_or_404(Categoria, pk=pk)
    
    if request.method == 'POST':
        # Verificar si tiene productos asociados
        if categoria.productos.exists():
            messages.error(
                request, 
                f'No se puede eliminar la categoría "{categoria.nombre}" porque tiene {categoria.productos.count()} producto(s) asociado(s).'
            )
            return redirect('categoria_lista')
        
        try:
            nombre = categoria.nombre
            categoria.delete()
            messages.success(request, f'Categoría "{nombre}" eliminada exitosamente.')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar la categoría: {str(e)}')
    
    return redirect('categoria_lista')


# ---------------------------
# Importación / Exportación
# ---------------------------

@login_required
def producto_descargar_plantilla(request):
    """
    Descarga una plantilla Excel para carga masiva de productos.
    Columnas:
    - Nombre, Descripción, Categoría, PrecioCompra, PrecioVenta, Estado(opcional: activo|inactivo)
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para gestionar productos.')
        return redirect('home')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'
    headers = ['Nombre', 'Descripción', 'Categoría', 'PrecioCompra', 'PrecioVenta', 'Estado']
    ws.append(headers)
    # Fila de ejemplo
    ws.append(['Galletas Chocochip', 'Paquete 12 unidades', 'Galletas', 10.50, 15.00, 'activo'])

    # Ancho básico de columnas
    widths = [30, 40, 20, 15, 15, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
    return response


@login_required
def producto_importar_excel(request):
    """
    Importa productos desde un archivo Excel (.xlsx). Crea la categoría si no existe.
    """
    if request.method != 'POST':
        return redirect('producto_listar')

    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permiso para gestionar productos.')
        return redirect('home')

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debes seleccionar un archivo Excel (.xlsx).')
        return redirect('producto_listar')

    try:
        wb = load_workbook(archivo)
        ws = wb.active
    except Exception as e:
        messages.error(request, f'No se pudo leer el archivo: {e}')
        return redirect('producto_listar')

    # Leer filas
    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    mapa = {h.lower().strip() if isinstance(h, str) else '': idx for idx, h in enumerate(encabezados)}
    requeridos = ['nombre', 'categoría', 'preciocompra', 'precioventa']
    faltantes = [r for r in requeridos if r not in mapa]
    if faltantes:
        messages.error(request, 'La plantilla no tiene las columnas requeridas: ' + ', '.join(faltantes))
        return redirect('producto_listar')

    creados = 0
    actualizados = 0
    errores = 0
    for row in ws.iter_rows(min_row=2):
        try:
            nombre = (row[mapa['nombre']].value or '').strip()
            if not nombre:
                continue  # fila vacía
            descripcion = (row[mapa.get('descripción', mapa.get('descripcion', -1))].value if ('descripción' in mapa or 'descripcion' in mapa) else '') or ''
            descripcion = descripcion.strip() if isinstance(descripcion, str) else str(descripcion or '')
            categoria_nombre = (row[mapa['categoría']].value or '').strip()
            precio_compra_val = row[mapa['preciocompra']].value
            precio_venta_val = row[mapa['precioventa']].value
            estado_val = (row[mapa.get('estado', -1)].value if 'estado' in mapa else 'activo') or 'activo'
            estado = str(estado_val).strip().lower()
            if estado not in ['activo', 'inactivo']:
                estado = 'activo'

            # Convertir precios
            def to_decimal(v):
                if v is None or v == '':
                    return None
                if isinstance(v, (int, float)):
                    return Decimal(str(v))
                try:
                    # Reemplazar coma por punto si viene con coma
                    return Decimal(str(v).replace(',', '.'))
                except (InvalidOperation, ValueError):
                    return None

            precio_compra = to_decimal(precio_compra_val)
            precio_venta = to_decimal(precio_venta_val)

            if precio_compra is None or precio_venta is None:
                raise ValueError('Precios inválidos')
            if not categoria_nombre:
                raise ValueError('Categoría requerida')

            categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)

            obj, creado = Producto.objects.get_or_create(
                nombre=nombre,
                defaults={
                    'descripcion': descripcion,
                    'categoria': categoria,
                    'precio_compra': precio_compra,
                    'precio_venta': precio_venta,
                    'estado': estado,
                }
            )
            if not creado:
                obj.descripcion = descripcion
                obj.categoria = categoria
                obj.precio_compra = precio_compra
                obj.precio_venta = precio_venta
                obj.estado = estado
                obj.save()
                actualizados += 1
            else:
                creados += 1
        except Exception as e:
            errores += 1
            logger.error('Error importando producto en fila %s: %s', row[0].row, e)

    if errores:
        messages.warning(request, f'Importación completa: {creados} creados, {actualizados} actualizados, {errores} con error.')
    else:
        messages.success(request, f'Importación completa: {creados} creados, {actualizados} actualizados.')

    return redirect('producto_listar')