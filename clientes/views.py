from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook, load_workbook
from django.db.models import Q
import logging
from .models import Cliente
from .forms import ClienteForm, ClienteFiltroForm, ClienteVendedorForm

logger = logging.getLogger(__name__)


@login_required
def cliente_listar(request):
    """
    Lista todos los clientes con filtros y paginación
    Acceso: admin y secretaria
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('home')
    
    # Obtener parámetros de filtro
    form_filtro = ClienteFiltroForm(request.GET)
    clientes = Cliente.objects.all()
    
    if form_filtro.is_valid():
        buscar = form_filtro.cleaned_data.get('buscar')
        activo = form_filtro.cleaned_data.get('activo')
        
        if buscar:
            clientes = clientes.filter(
                Q(nombre__icontains=buscar) | 
                Q(nit__icontains=buscar) |
                Q(nombre_contacto__icontains=buscar)
            )
        
        if activo == '1':
            clientes = clientes.filter(activo=True)
        elif activo == '0':
            clientes = clientes.filter(activo=False)
    
    clientes = clientes.order_by('nombre')
    
    # Paginación
    paginator = Paginator(clientes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form_filtro': form_filtro,
    }
    
    return render(request, 'clientes/cliente_list.html', context)


@login_required
def cliente_crear(request):
    """
    Crea un nuevo cliente
    Acceso: admin y secretaria
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('home')
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Cliente creado exitosamente!')
            return redirect('cliente_listar')
        else:
            messages.error(request, 'Error al crear el cliente. Verifique los datos.')
            try:
                logger.warning('cliente_crear POST keys: %s', list(request.POST.keys()))
                logger.error('cliente_crear form errors: %s', form.errors.as_json())
            except Exception:
                pass
    else:
        form = ClienteForm()
    
    context = {'form': form}
    return render(request, 'clientes/cliente_form.html', context)


@login_required
def cliente_editar(request, pk):
    """
    Edita un cliente existente
    Acceso: admin y secretaria
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('home')
    
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Cliente actualizado exitosamente!')
            return redirect('cliente_listar')
        else:
            messages.error(request, 'Error al actualizar el cliente. Verifique los datos.')
    else:
        form = ClienteForm(instance=cliente)
    
    context = {'form': form, 'cliente': cliente}
    return render(request, 'clientes/cliente_form.html', context)


@login_required
def cliente_eliminar(request, pk):
    """
    Desactiva un cliente (soft delete)
    Acceso: admin y secretaria
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('home')
    
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = False
    cliente.save()
    
    messages.success(request, 'Cliente desactivado exitosamente.')
    return redirect('cliente_listar')


@login_required
def cliente_activar(request, pk):
    """
    Activa un cliente
    Acceso: admin y secretaria
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('home')
    
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = True
    cliente.save()
    
    messages.success(request, 'Cliente activado exitosamente.')
    return redirect('cliente_listar')


@login_required
def vendedor_crear_cliente(request):
    """
    Permite al vendedor crear un cliente nuevo que se auto-asigna a su ruta activa
    El cliente aparece inmediatamente en su planificación del día
    Acceso: solo vendedores
    """
    if not request.user.es_vendedor:
        messages.error(request, 'Esta sección es solo para vendedores.')
        return redirect('login')
    
    from asignaciones.models import Asignacion
    from rutas.models import RutaDetalle
    from planificacion.models import Planificacion
    from datetime import date
    
    # En este proyecto el vendedor es el propio Usuario con rol 'vendedor'
    vendedor = request.user
    
    # Verificar que tenga asignación activa
    asignaciones_activas = []
    for asig in Asignacion.objects.filter(vendedor=vendedor):
        if asig.esta_activa:
            asignaciones_activas.append(asig)
    
    asignacion = asignaciones_activas[0] if asignaciones_activas else None
    
    if not asignacion:
        messages.warning(request, 'No tienes una ruta asignada. Contacta con el administrador.')
        return redirect('planificacion_vendedor_dia')
    
    if request.method == 'POST':
        form = ClienteVendedorForm(request.POST)
        if form.is_valid():
            # Crear cliente
            cliente = form.save()
            
            # Obtener el máximo orden_visita de la ruta
            max_orden = RutaDetalle.objects.filter(
                ruta=asignacion.ruta
            ).count() + 1
            
            # Asignar a la ruta del vendedor (crear RutaDetalle)
            ruta_detalle = RutaDetalle.objects.create(
                ruta=asignacion.ruta,
                cliente=cliente,
                orden_visita=max_orden,
                activo=True
            )
            
            # Crear planificación para hoy con tipo 'no_planificado'
            hoy = date.today()
            planificacion = Planificacion.objects.create(
                asignacion=asignacion,
                ruta_detalle=ruta_detalle,
                fecha=hoy,
                tipo='no_planificado'
            )
            
            messages.success(request, f'¡Cliente "{cliente.nombre}" registrado y agregado a tu ruta!')
            return redirect('planificacion_vendedor_dia')
        else:
            messages.error(request, 'Error al crear el cliente. Verifique los datos.')
            try:
                logger.warning('vendedor_crear_cliente POST keys: %s', list(request.POST.keys()))
                logger.error('vendedor_crear_cliente form errors: %s', form.errors.as_json())
            except Exception:
                pass
    else:
        form = ClienteVendedorForm()
    
    context = {
        'form': form,
        'asignacion': asignacion,
        'es_vendedor': True
    }
    return render(request, 'clientes/vendedor_crear_cliente.html', context)


# ---------------------------
# Importación / Exportación
# ---------------------------

@login_required
def cliente_descargar_plantilla(request):
    """
    Descarga una plantilla Excel para carga masiva de clientes.
    Columnas:
    - NIT, Nombre, NombreContacto, Correo, Telefono, Direccion, ReferenciaUbicacion, Latitud, Longitud, Activo
    """
    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('home')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    headers = ['NIT', 'Nombre', 'NombreContacto', 'Correo', 'Telefono', 'Direccion',
               'ReferenciaUbicacion', 'Latitud', 'Longitud', 'Activo']
    ws.append(headers)
    ws.append(['1234567-8', 'Tienda La Esquina', 'Juan Perez', 'tienda@example.com', '5555-5555',
               '1a calle 2-34 zona 1', 'Frente a parque', 14.6345, -90.5069, 1])

    widths = [15, 28, 22, 28, 14, 35, 28, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_clientes.xlsx"'
    return response


@login_required
def cliente_importar_excel(request):
    """
    Importa clientes desde Excel (.xlsx). Si el NIT existe, se actualiza; si no, se crea.
    """
    if request.method != 'POST':
        return redirect('cliente_listar')

    if not request.user.puede_gestionar_rutas:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('home')

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debes seleccionar un archivo Excel (.xlsx).')
        return redirect('cliente_listar')

    try:
        wb = load_workbook(archivo)
        ws = wb.active
    except Exception as e:
        messages.error(request, f'No se pudo leer el archivo: {e}')
        return redirect('cliente_listar')

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    mapa = {str(h).lower().strip(): idx for idx, h in enumerate(headers)}
    requeridos = ['nit', 'nombre', 'telefono', 'direccion']
    faltantes = [r for r in requeridos if r not in mapa]
    if faltantes:
        messages.error(request, 'La plantilla no tiene las columnas requeridas: ' + ', '.join(faltantes))
        return redirect('cliente_listar')

    creados, actualizados, errores = 0, 0, 0
    for row in ws.iter_rows(min_row=2):
        try:
            nit = (row[mapa['nit']].value or '').strip()
            if not nit:
                continue
            nombre = (row[mapa['nombre']].value or '').strip()
            nombre_contacto = (row[mapa.get('nombrecontacto', -1)].value if 'nombrecontacto' in mapa else '') or ''
            correo = (row[mapa.get('correo', -1)].value if 'correo' in mapa else '') or ''
            telefono = (row[mapa['telefono']].value or '').strip()
            direccion = (row[mapa['direccion']].value or '').strip()
            ref = (row[mapa.get('referenciaubicacion', -1)].value if 'referenciaubicacion' in mapa else '') or ''
            lat = row[mapa.get('latitud', -1)].value if 'latitud' in mapa else None
            lon = row[mapa.get('longitud', -1)].value if 'longitud' in mapa else None
            activo_val = row[mapa.get('activo', -1)].value if 'activo' in mapa else 1

            def to_decimal(v):
                if v in (None, ''):
                    return None
                try:
                    return Decimal(str(v))
                except Exception:
                    return None

            latitud = to_decimal(lat)
            longitud = to_decimal(lon)
            activo = True
            if isinstance(activo_val, str):
                activo = activo_val.strip().lower() in ['1', 'true', 'si', 'sí', 'activo']
            elif isinstance(activo_val, (int, float)):
                activo = int(activo_val) != 0

            obj, creado = Cliente.objects.get_or_create(
                nit=nit,
                defaults={
                    'nombre': nombre,
                    'nombre_contacto': nombre_contacto,
                    'correo': correo or None,
                    'telefono': telefono,
                    'direccion': direccion,
                    'referencia_ubicacion': ref,
                    'latitud': latitud,
                    'longitud': longitud,
                    'activo': activo,
                }
            )
            if not creado:
                obj.nombre = nombre or obj.nombre
                obj.nombre_contacto = nombre_contacto
                obj.correo = correo or None
                obj.telefono = telefono or obj.telefono
                obj.direccion = direccion or obj.direccion
                obj.referencia_ubicacion = ref
                obj.latitud = latitud
                obj.longitud = longitud
                obj.activo = activo
                obj.save()
                actualizados += 1
            else:
                creados += 1
        except Exception as e:
            errores += 1
            logger.error('Error importando cliente en fila %s: %s', row[0].row, e)

    if errores:
        messages.warning(request, f'Importación completa: {creados} creados, {actualizados} actualizados, {errores} con error.')
    else:
        messages.success(request, f'Importación completa: {creados} creados, {actualizados} actualizados.')
    return redirect('cliente_listar')
